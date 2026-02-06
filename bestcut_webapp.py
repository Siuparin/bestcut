# bestcut_webapp.py
# Versione con taglio parziale - dice cosa si può fare e cosa manca

import streamlit as st
from dataclasses import dataclass
from typing import List, Tuple, Dict
import copy
from itertools import combinations
from datetime import datetime
import pandas as pd
from io import BytesIO

# Importazione openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_DISPONIBILE = True
except ImportError:
    EXCEL_DISPONIBILE = False

# Configurazione pagina
st.set_page_config(
    page_title="BestCut - Taglio Parziale",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizzato
st.markdown("""
<style>
    .main-header {
        font-size: 3rem;
        font-weight: bold;
        color: #2196F3;
        text-align: center;
        margin-bottom: 0;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .partial-box {
        background-color: #E3F2FD;
        color: #1565C0;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #2196F3;
        margin: 1rem 0;
    }
    .success-box {
        background-color: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #28a745;
    }
    .error-box {
        background-color: #f8d7da;
        color: #721c24;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #dc3545;
    }
    .missing-box {
        background-color: #ffebee;
        color: #c62828;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #f44336;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)

@dataclass
class Spezzone:
    lunghezza: float
    id: int
    
@dataclass
class TaglioRichiesto:
    lunghezza: float
    quantita: int
    
@dataclass
class PianoTaglio:
    spezzone_id: int
    spezzone_lunghezza: float
    tagli: List[float]
    scarto: float

@dataclass
class RisultatoCalcolo:
    piani: List[PianoTaglio]
    scarto_totale: float
    completato: bool  # True = tutto fatto, False = parziale
    tagli_fatti: Dict[float, int]  # misura -> quantità fatta
    tagli_mancanti: Dict[float, int]  # misura -> quantità mancante
    spezzoni_usati: int
    spezzoni_totali: int

class OttimizzatoreTagli:
    def __init__(self, soglia_scarto: float = 0.3):
        self.soglia_scarto = soglia_scarto
        
    def calcola_ottimale(self, spezzoni: List[Spezzone], richieste: List[TaglioRichiesto]) -> RisultatoCalcolo:
        """
        Calcola il piano di taglio.
        Se non basta il materiale, fa quello che può e indica cosa manca.
        """
        # Crea lista di tutti i tagli necessari con la loro misura
        tagli_necessari = []
        for richiesta in richieste:
            for _ in range(richiesta.quantita):
                tagli_necessari.append(richiesta.lunghezza)
        
        # Ordina per lunghezza decrescente (i più grandi prima)
        tagli_necessari.sort(reverse=True)
        
        # Copia per tenere traccia di cosa riusciamo a fare
        tagli_originali = tagli_necessari.copy()
        
        # Ordina spezzoni dal più grande al più piccolo
        spezzoni_ordinati = sorted(spezzoni, key=lambda x: x.lunghezza, reverse=True)
        spezzoni_work = copy.deepcopy(spezzoni_ordinati)
        tagli_rimanenti = tagli_necessari.copy()
        piani = []
        
        while tagli_rimanenti and spezzoni_work:
            spezzone_corrente = spezzoni_work[0]
            tagli_da_tagliare = []
            tagli_temp = tagli_rimanenti.copy()
            
            # Prova a infilare tagli dal più grande al più piccolo
            for taglio in tagli_temp:
                if sum(tagli_da_tagliare) + taglio <= spezzone_corrente.lunghezza:
                    tagli_da_tagliare.append(taglio)
                    tagli_rimanenti.remove(taglio)
            
            if tagli_da_tagliare:
                scarto = spezzone_corrente.lunghezza - sum(tagli_da_tagliare)
                piani.append(PianoTaglio(
                    spezzone_id=spezzone_corrente.id,
                    spezzone_lunghezza=spezzone_corrente.lunghezza,
                    tagli=tagli_da_tagliare,
                    scarto=scarto
                ))
            
            spezzoni_work.pop(0)
        
        # Calcola cosa è stato fatto e cosa manca
        tagli_fatti = {}
        tagli_mancanti = {}
        
        for piano in piani:
            for taglio in piano.tagli:
                tagli_fatti[taglio] = tagli_fatti.get(taglio, 0) + 1
        
        # Per ogni richiesta originale, calcola quanti ne mancano
        for richiesta in richieste:
            fatti = tagli_fatti.get(richiesta.lunghezza, 0)
            if fatti < richiesta.quantita:
                tagli_mancanti[richiesta.lunghezza] = richiesta.quantita - fatti
        
        scarto_totale = sum(p.scarto for p in piani)
        completato = len(tagli_rimanenti) == 0
        
        return RisultatoCalcolo(
            piani=piani,
            scarto_totale=scarto_totale,
            completato=completato,
            tagli_fatti=tagli_fatti,
            tagli_mancanti=tagli_mancanti,
            spezzoni_usati=len(piani),
            spezzoni_totali=len(spezzoni)
        )

def crea_excel_download(spezzoni, richieste, risultato, soglia):
    """Crea file Excel in memoria per il download"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Piano Taglio"
    
    header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    subheader_font = Font(name='Calibri', size=12, bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:E1')
    ws['A1'] = "PIANO DI TAGLIO TUBI"
    ws['A1'].font = Font(size=18, bold=True, color="2196F3")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Generato: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True)
    
    # Stato completamento
    row = 4
    if risultato.completato:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "✅ TAGLIO COMPLETATO - Tutti i pezzi realizzabili"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="4CAF50")
    else:
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "⚠️ TAGLIO PARZIALE - Materiali insufficienti"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FF9800")
    
    # Spezzoni disponibili
    row = 6
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SPEZZONI DISPONIBILI"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    for col, title in [('A', 'ID'), ('B', 'Lunghezza (m)'), ('C', 'Lunghezza (cm)')]:
        ws[f'{col}{row}'] = title
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].border = border
    
    for spezzone in spezzoni:
        row += 1
        ws[f'A{row}'] = spezzone.id
        ws[f'B{row}'] = spezzone.lunghezza
        ws[f'C{row}'] = spezzone.lunghezza * 100
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
    
    # Tagli richiesti vs fatti
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "RIEPILOGO TAGLI"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 1
    headers = [('A', 'Misura (m)'), ('B', 'Misura (cm)'), ('C', 'Quantita'), ('D', 'Totale (m)'), ('E', 'Tubo mancante (m)')]
    for col, title in headers:
        ws[f'{col}{row}'] = title
        ws[f'{col}{row}'].font = subheader_font
        ws[f'{col}{row}'].border = border
    
    for richiesta in richieste:
        row += 1
        ws[f'A{row}'] = richiesta.lunghezza
        ws[f'B{row}'] = richiesta.lunghezza * 100
        ws[f'C{row}'] = richiesta.quantita
        ws[f'D{row}'] = richiesta.lunghezza * richiesta.quantita
        
        # NUOVO: Calcola tubo mancante
        fatti = risultato.tagli_fatti.get(richiesta.lunghezza, 0)
        mancanti = risultato.tagli_mancanti.get(richiesta.lunghezza, 0)
        tubo_mancante = mancanti * richiesta.lunghezza if mancanti > 0 else 0
        ws[f'E{row}'] = tubo_mancante if tubo_mancante > 0 else "-"
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = border

    # AGGIUNTO: Riga vuota per separazione
    row += 1
    
    # Riga totale tubo mancante
    row += 1
    ws[f'A{row}'] = "TOTALE TUBO MANCANTE:"
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:E{row}')
    totale_mancante = sum(misura * qty for misura, qty in risultato.tagli_mancanti.items())
    ws[f'B{row}'] = f"{totale_mancante:.2f} m" if totale_mancante > 0 else "0 m"
    ws[f'B{row}'].font = Font(bold=True, color="f44336" if totale_mancante > 0 else "4CAF50")
    for col in ['A', 'B']:
        ws[f'{col}{row}'].border = border
    
    # Piano di taglio dettagliato
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "PIANO DI TAGLIO DETTAGLIATO"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    for piano in risultato.piani:
        row += 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Spezzone #{piano.spezzone_id} ({piano.spezzone_lunghezza:.3f}m)"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="E3F2FD", fill_type="solid")
        
        row += 1
        headers = [('A', 'N°'), ('B', 'Misura (m)'), ('C', 'Misura (cm)'), 
                  ('D', 'Inizio (m)'), ('E', 'Fine (m)')]
        for col, title in headers:
            ws[f'{col}{row}'] = title
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].fill = PatternFill(start_color="F5F5F5", fill_type="solid")
        
        posizione = 0.0
        for i, taglio in enumerate(piano.tagli, 1):
            row += 1
            ws[f'A{row}'] = i
            ws[f'B{row}'] = taglio
            ws[f'C{row}'] = taglio * 100
            ws[f'D{row}'] = posizione
            ws[f'E{row}'] = posizione + taglio
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = border
            posizione += taglio
        
        row += 1
        ws[f'A{row}'] = "SCARTO"
        color = "4CAF50" if piano.scarto <= soglia else "f44336"
        ws[f'A{row}'].font = Font(bold=True, color=color)
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = piano.scarto
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'D{row}'] = "OTTIMALE" if piano.scarto <= soglia else "DA RIUTILIZZARE"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = border
        row += 1
    
    for i, w in enumerate([15, 15, 15, 18, 18], 1):
        ws.column_dimensions[chr(64+i)].width = w
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def main():
    # Header
    st.markdown('<p class="main-header">🔧 BestCut v3.2</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Ottimizzatore con supporto TAGLIO PARZIALE</p>', unsafe_allow_html=True)
    
    # Info
    st.markdown("""
    <div class="partial-box">
        <strong>🆕 NOVITÀ:</strong> Se i tubi non bastano, il programma ti dice cosa riesci a fare 
        con quello che hai e quanto ne manca!
    </div>
    """, unsafe_allow_html=True)
    
    # Inizializza session state
    if 'spezzoni' not in st.session_state:
        st.session_state.spezzoni = []
        st.session_state.prossimo_id = 1
        st.session_state.risultato = None
        st.session_state.richieste = None
        st.session_state.soglia = 0.3
    
    # Layout a colonne
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("📦 Spezzoni Disponibili")
        
        st.info("💡 Inserisci prima i tubi PIÙ GRANDI")
        
        nuovo_spezzone = st.number_input(
            "Lunghezza spezzone (metri)",
            min_value=0.0,
            value=6.0,
            step=0.1,
            format="%.2f",
            key="input_spezzone"
        )
        
        if st.button("➕ Aggiungi Spezzone", use_container_width=True):
            if nuovo_spezzone > 0:
                st.session_state.spezzoni.append(
                    Spezzone(nuovo_spezzone, st.session_state.prossimo_id)
                )
                st.session_state.spezzoni.sort(key=lambda x: x.lunghezza, reverse=True)
                for i, s in enumerate(st.session_state.spezzoni, 1):
                    s.id = i
                st.session_state.prossimo_id = len(st.session_state.spezzoni) + 1
                st.success(f"✅ Aggiunto: {nuovo_spezzone:.2f}m")
                st.rerun()
            else:
                st.error("❌ Lunghezza non valida")
        
        if st.session_state.spezzoni:
            data = [{"ID": s.id, "Lunghezza (m)": f"{s.lunghezza:.2f}", "Lunghezza (cm)": f"{s.lunghezza*100:.0f}"} 
                   for s in st.session_state.spezzoni]
            df = pd.DataFrame(data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            id_da_rimuovere = st.selectbox(
                "Seleziona da rimuovere",
                options=[s.id for s in st.session_state.spezzoni],
                format_func=lambda x: f"ID {x} - {next(s.lunghezza for s in st.session_state.spezzoni if s.id == x):.2f}m"
            )
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("🗑️ Rimuovi", use_container_width=True):
                    st.session_state.spezzoni = [s for s in st.session_state.spezzoni if s.id != id_da_rimuovere]
                    for i, s in enumerate(st.session_state.spezzoni, 1):
                        s.id = i
                    st.session_state.prossimo_id = len(st.session_state.spezzoni) + 1
                    st.success("✅ Rimosso!")
                    st.rerun()
            with col_btn2:
                if st.button("🗑️🗑️ Tutti", use_container_width=True):
                    st.session_state.spezzoni = []
                    st.session_state.prossimo_id = 1
                    st.success("✅ Tutti rimossi!")
                    st.rerun()
        else:
            st.warning("⚠️ Nessuno spezzone inserito")
    
    with col2:
        st.subheader("✂️ Tagli Richiesti")
        
        st.session_state.soglia = st.number_input(
            "Soglia scarto (metri)",
            min_value=0.0,
            value=0.3,
            step=0.05,
            format="%.2f"
        )
        
        st.markdown("---")
        richieste = []
        
        for i in range(5):
            cols = st.columns([1, 2, 2])
            with cols[0]:
                st.markdown(f"**#{i+1}**")
            with cols[1]:
                misura = st.number_input(
                    f"M{i+1}", min_value=0.0,
                    value=3.2 if i == 0 else (0.5 if i == 1 else 0.0),
                    step=0.1, format="%.2f",
                    key=f"misura_{i}", label_visibility="collapsed"
                )
            with cols[2]:
                qty = st.number_input(
                    f"Q{i+1}", min_value=0,
                    value=1 if i == 0 else (5 if i == 1 else 0),
                    step=1, key=f"qty_{i}", label_visibility="collapsed"
                )
            
            if misura > 0 and qty > 0:
                richieste.append(TaglioRichiesto(misura, qty))
        
        richieste.sort(key=lambda x: x.lunghezza, reverse=True)
        st.session_state.richieste = richieste
        
        st.markdown("---")
        st.write(f"**{len(richieste)} tipi di tagli configurati**")
    
    # Bottone calcola
    st.markdown("---")
    col_center = st.columns([1, 2, 1])
    with col_center[1]:
        if st.button("🚀 CALCOLA (anche parziale)", use_container_width=True, type="primary"):
            if not st.session_state.spezzoni:
                st.error("❌ Aggiungi almeno uno spezzone!")
            elif not richieste:
                st.error("❌ Inserisci almeno un taglio!")
            else:
                with st.spinner("⏳ Calcolo in corso..."):
                    ottim = OttimizzatoreTagli(st.session_state.soglia)
                    risultato = ottim.calcola_ottimale(
                        copy.deepcopy(st.session_state.spezzoni), 
                        richieste
                    )
                    st.session_state.risultato = risultato
                
                if risultato.completato:
                    st.success("✅ TAGLIO COMPLETATO! Tutti i pezzi realizzabili")
                else:
                    st.warning("⚠️ TAGLIO PARZIALE - Materiali insufficienti")
    
    # Risultati
    if st.session_state.risultato:
        st.markdown("---")
        
        risultato = st.session_state.risultato
        richieste = st.session_state.richieste
        
        # Box stato
        if risultato.completato:
            st.markdown('<div class="success-box">✅ <strong>COMPLETATO!</strong> Tutti i tagli sono realizzabili con gli spezzoni disponibili.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="partial-box">⚠️ <strong>PARZIALE!</strong> Con gli spezzoni disponibili riesci a fare solo una parte dei tagli richiesti.</div>', unsafe_allow_html=True)
        
        # Metriche
        scarto_tot = risultato.scarto_totale
        efficienza = (1 - scarto_tot/sum(p.spezzone_lunghezza for p in risultato.piani))*100 if risultato.piani else 0
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
            st.metric("Spezzoni usati", f"{risultato.spezzoni_usati}/{risultato.spezzoni_totali}")
        with col_m2:
            st.metric("Scarto totale", f"{scarto_tot:.3f}m")
        with col_m3:
            st.metric("Efficienza", f"{efficienza:.1f}%")
        with col_m4:
            risparmiati = risultato.spezzoni_totali - risultato.spezzoni_usati
            st.metric("💰 Risparmiati", risparmiati if risparmiati > 0 else 0)
        
        # Tabella riepilogo: Richiesti vs Fatti vs Mancanti
        st.subheader("📊 Riepilogo Tagli")
        
        data_riep = []
        for rich in richieste:
            fatti = risultato.tagli_fatti.get(rich.lunghezza, 0)
            mancanti = risultato.tagli_mancanti.get(rich.lunghezza, 0)
            
            data_riep.append({
                "Misura": f"{rich.lunghezza:.2f}m",
                "Richiesti": rich.quantita,
                "✅ Fatti": fatti,
                "❌ Mancanti": mancanti if mancanti > 0 else "-",
                "Stato": "🟢 OK" if mancanti == 0 else f"🟡 Mancano {mancanti}"
            })
        
        df_riep = pd.DataFrame(data_riep)
        st.dataframe(df_riep, use_container_width=True, hide_index=True)
        
        # Avviso se manca qualcosa
        if not risultato.completato:
            st.markdown("---")
            st.subheader("❌ Tagli Mancanti")
            
            for misura, qty in risultato.tagli_mancanti.items():
                st.markdown(f"""
                <div class="missing-box">
                    <strong>{misura:.2f}m</strong>: mancano <strong>{qty} pezzi</strong><br>
                    <small>Servono altri {misura * qty:.2f}m di tubo per completare</small>
                </div>
                """, unsafe_allow_html=True)
            
            totale_mancante = sum(misura * qty for misura, qty in risultato.tagli_mancanti.items())
            st.info(f"💡 In totale mancano {totale_mancante:.2f}m di tubo per completare tutti i tagli")
        
        # Dettaglio piano di taglio
        st.markdown("---")
        st.subheader("🔧 Piano di Taglio Dettagliato")
        
        for piano in risultato.piani:
            with st.expander(f"Spezzone #{piano.spezzone_id} ({piano.spezzone_lunghezza:.3f}m)"):
                data_tagli = []
                pos = 0.0
                for i, taglio in enumerate(piano.tagli, 1):
                    data_tagli.append({
                        "N°": i,
                        "Misura (m)": f"{taglio:.3f}",
                        "Misura (cm)": f"{taglio*100:.1f}",
                        "Inizio": f"{pos:.3f}m",
                        "Fine": f"{pos+taglio:.3f}m"
                    })
                    pos += taglio
                
                df_tagli = pd.DataFrame(data_tagli)
                st.dataframe(df_tagli, use_container_width=True, hide_index=True)
                
                if piano.scarto <= st.session_state.soglia:
                    st.success(f"✅ Scarto: {piano.scarto:.3f}m - OTTIMALE")
                else:
                    st.warning(f"⚠️ Scarto: {piano.scarto:.3f}m - DA RIUTILIZZARE")
        
        # Download Excel
        st.markdown("---")
        if EXCEL_DISPONIBILE:
            excel_buffer = crea_excel_download(
                st.session_state.spezzoni,
                richieste,
                risultato,
                st.session_state.soglia
            )
            
            col_dl1, col_dl2, col_dl3 = st.columns([1, 2, 1])
            with col_dl2:
                st.download_button(
                    label="📥 Scarica Report Excel",
                    data=excel_buffer,
                    file_name=f"BestCut_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.error("⚠️ openpyxl non installato")

if __name__ == "__main__":
    main()
